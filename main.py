from openpyxl import load_workbook
from datetime import datetime, timedelta
from datetime import time
from openpyxl.styles import PatternFill
import numpy as np

wb1 = load_workbook('wykaz zajęć 4-24 maj 2020r.xlsx')
ws1 = wb1['Arkusz1'] # Work Sheet

wb2 = load_workbook('AttendeeReport_WSIZ-EU_2020-05-08_2020-05-10.xlsx')
ws2 = wb2['Sheet1'] # Work Sheet

x = 2
position = []
surname1 = ""
surname2 = ""
name1 = ""
name2 = ""

print ("Start")
while x <= 3569: #ADJUST
    arr_start = []
    arr_end = []

    # Gathering times of classes that should take place 
    while True:
        #If classe is cancelled then go to next
        #if ws1.cell(row = x, column = 11).value == "Odwołane":
        #    x += 1
        surname1 = ws1.cell(row = x, column = 3).value
        surname2 = ws1.cell(row = x+1, column = 3).value
        name1 = ws1.cell(row = x, column = 4).value
        name2 = ws1.cell(row = x+1, column = 4).value

        if surname1 == surname2 and name1 == name2:

            #Start/End time from the Dean's Office list
            arr_start.append(ws1.cell(row = x, column = 1).value)
            arr_end.append(ws1.cell(row = x, column = 2).value)

        #If that's the last record under current surname add it to array
        elif x > 1 and surname1 == ws1.cell(row = x-1, column = 3).value and name1 == ws1.cell(row = x-1, column = 4).value:

            arr_start.append(ws1.cell(row = x, column = 1).value)
            arr_end.append(ws1.cell(row = x, column = 2).value)
            x+=1
            break

        #If that's the sole record of current surname -> add to array
        else:
            arr_start.append(ws1.cell(row = x, column = 1).value)
            arr_end.append(ws1.cell(row = x, column = 2).value)
            x+=1
            break
        x+=1

    #Filtering/Deleting same dates  
    o = 0
    z = 0
    while z < len(arr_start)-1:
        if o == 1:
            z = 0
            o = 0
        if arr_start[z] == arr_start[z+1] and arr_end[z] == arr_end[z+1]:
            del arr_start[z]
            del arr_end[z]
            o = 1
        z+=1

    #for u in range(len(arr_start)):
    #    print (arr_start[u], arr_end[u])
    
    #Gathering second array with classes that took place on this surname
    arr2_start = []
    arr2_end = []
    i = 0
    for y in range(2, 33252): #ADJUST last record number and below columns
        if ws2.cell(row = y, column = 6).value == "Moderator":
            surname2 = ws2.cell(row = y, column = 5).value
            duration = ws2.cell(row = y, column = 4).value
            #If current record contains full name of the professor
            contains = surname1 in surname2 and name1 in surname2
            if contains:
                c = time(00, 40, 00)
                etalon = datetime(2020, 3, 29)
                if isinstance(duration, time):
                    if duration > c:
                        #Standart Time
                        if ws2.cell(row = y, column = 3).value < etalon:
                            arr2_start.append(ws2.cell(row = y, column = 2).value + timedelta(hours=1))
                            arr2_end.append(ws2.cell(row = y, column = 3).value + timedelta(hours=1))
                        #Summer Time
                        else:
                            arr2_start.append(ws2.cell(row = y, column = 2).value + timedelta(hours=2))
                            arr2_end.append(ws2.cell(row = y, column = 3).value + timedelta(hours=2))
    h = 0
    #Looping simultaneously through both arrays
    while h < len(arr_start) and h < len(arr2_start):
        #If the record from Dean's office within the range of the session that took place
        q = 2
        #Loop through Dean's list to find all matching records to color green, change status to 'OK'
        while True:
            # When full name, start and end times match -> this is our record 
            if ws1.cell(row = q, column = 3).value == surname1 and ws1.cell(row = q, column = 4).value == name1 and ws1.cell(row = q, column = 1).value == arr_start[h] and ws1.cell(row = q, column = 2).value == arr_end[h]:
                if (ws1.cell(row = q, column = 1).value >= (arr2_start[h] - timedelta(minutes = 10))) and (ws1.cell(row = q, column = 2).value <= (arr2_end[h] + timedelta(minutes = 35))):
                    ws1.cell(row = q, column = 15).value = 'OK'
                    for col_range in range(1, 16):
                        cell_title = ws1.cell(q, col_range)
                        cell_title.fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
            q+=1
            # BREAK
            if q == 3569: #ADJUST
                break
        h+=1

print ("END")
wb1.save('Result.xlsx')